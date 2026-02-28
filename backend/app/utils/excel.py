from collections.abc import Callable
from datetime import date, datetime, time
from decimal import Decimal
from enum import Enum
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def _format_value(value):
    """Convert a value to an Excel-friendly format."""
    if value is None:
        return ""
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def _resolve_value(item, field: str | Callable):
    """Resolve a column value from an item using a field name or callable."""
    if callable(field):
        return field(item)
    return getattr(item, field, None)


def generate_excel(
    items: list,
    columns: list[tuple[str | Callable, str]],
    sheet_name: str = "Sheet1",
) -> BytesIO:
    """Generate an Excel file from a list of model instances.

    Args:
        items: List of SQLAlchemy model instances.
        columns: List of (field_or_callable, header_label) tuples.
                 field_or_callable can be a string attribute name or a
                 callable that takes an item and returns the value.
        sheet_name: Name for the worksheet.

    Returns:
        BytesIO buffer containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    bold = Font(bold=True)

    # Write headers
    for col_idx, (_, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Write data rows
    for row_idx, item in enumerate(items, start=2):
        for col_idx, (field, _) in enumerate(columns, start=1):
            value = _resolve_value(item, field)
            ws.cell(row=row_idx, column=col_idx, value=_format_value(value))

    # Auto-width columns
    for col_idx, (_, header) in enumerate(columns, start=1):
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
